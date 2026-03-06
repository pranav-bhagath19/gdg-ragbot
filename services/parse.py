"""
services/parse.py - Document parsers for PDF, Excel, and DOCX files.
Place this file at: rag-chatbot/services/parse.py
"""
import io
from pathlib import Path
from typing import List, Dict, Any
from loguru import logger


def parse_pdf(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Parse a PDF file and extract text with page metadata.
    Returns a list of dicts: {text, page, source}
    """
    import pdfplumber

    chunks = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                text = page.extract_text()
                if text and text.strip():
                    chunks.append({
                        "text": text.strip(),
                        "page": page_num,
                        "source": filename,
                        "type": "pdf"
                    })
        logger.info(f"Parsed PDF '{filename}': {len(chunks)} pages with text")
    except Exception as e:
        logger.error(f"Error parsing PDF '{filename}': {e}")
        raise ValueError(f"Failed to parse PDF: {str(e)}")
    return chunks


def parse_xlsx(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Parse an Excel file, converting each row into a natural language string.
    Returns a list of dicts: {text, page (sheet), source}
    """
    import openpyxl

    chunks = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            rows_data = []

            # Fields to skip (noisy, not useful for Q&A)
            _skip = {"id", "imageurl", "linkedinurl", "bgcolor", "logo", "rank",
                     "dept_rank", "dept_logo"}

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                # Skip completely empty rows
                if all(cell is None for cell in row):
                    continue

                if row_idx == 0:
                    # First row is headers — filter out noisy columns
                    all_headers = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(row)]
                    keep_cols = []
                    for i, h in enumerate(all_headers):
                        if h.lower().strip() not in _skip:
                            keep_cols.append(i)
                    headers = [all_headers[i] for i in keep_cols]
                else:
                    # Extract only kept columns, skip URL values
                    vals = []
                    for i in keep_cols:
                        cell = row[i] if i < len(row) else None
                        if cell is not None:
                            val = str(cell).strip()
                            if val.startswith(("http://", "https://")):
                                val = ""
                        else:
                            val = ""
                        vals.append(val)
                    if any(v for v in vals):
                        rows_data.append(vals)

            if rows_data:
                # Build compact table: headers once, then pipe-separated rows
                # This is much more token-efficient than repeating headers per row
                header_line = " | ".join(headers)
                batch_size = 10
                for batch_start in range(0, len(rows_data), batch_size):
                    batch = rows_data[batch_start: batch_start + batch_size]
                    row_lines = [" | ".join(v for v in row) for row in batch]
                    text = f"Sheet: {sheet_name}\n{header_line}\n" + "\n".join(row_lines)
                    chunks.append({
                        "text": text,
                        "page": f"{sheet_name}",
                        "source": filename,
                        "type": "xlsx"
                    })

        logger.info(f"Parsed XLSX '{filename}': {len(chunks)} chunks")
    except Exception as e:
        logger.error(f"Error parsing XLSX '{filename}': {e}")
        raise ValueError(f"Failed to parse Excel file: {str(e)}")
    return chunks


def parse_docx(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Parse a DOCX file, extracting paragraphs grouped by heading.
    Returns a list of dicts: {text, page (heading/section), source}
    """
    from docx import Document
    from docx.oxml.ns import qn

    chunks = []
    try:
        doc = Document(io.BytesIO(file_bytes))
        current_heading = "Introduction"
        current_paragraphs = []

        def flush_section():
            nonlocal current_paragraphs
            text = "\n".join(current_paragraphs).strip()
            if text:
                chunks.append({
                    "text": f"{current_heading}\n{text}",
                    "page": current_heading,
                    "source": filename,
                    "type": "docx"
                })
            current_paragraphs = []

        for para in doc.paragraphs:
            style_name = para.style.name if para.style else ""
            text = para.text.strip()
            if not text:
                continue

            if "Heading" in style_name:
                flush_section()
                current_heading = text
            else:
                current_paragraphs.append(text)

        flush_section()

        logger.info(f"Parsed DOCX '{filename}': {len(chunks)} sections")
    except Exception as e:
        logger.error(f"Error parsing DOCX '{filename}': {e}")
        raise ValueError(f"Failed to parse DOCX: {str(e)}")
    return chunks


def parse_file(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Auto-detect file type and parse accordingly.
    Supports: .pdf, .xlsx, .xls, .csv, .docx
    """
    ext = Path(filename).suffix.lower()

    if ext == ".pdf":
        return parse_pdf(file_bytes, filename)
    elif ext in (".xlsx", ".xls"):
        return parse_xlsx(file_bytes, filename)
    elif ext == ".docx":
        return parse_docx(file_bytes, filename)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Supported: .pdf, .xlsx, .xls, .docx")
